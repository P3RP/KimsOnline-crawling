# -*- encoding:utf-8 -*-

import os
import time
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import json
import re
import logging.handlers
from tqdm import tqdm


def my_debug():
    input("정지")
    driver.quit()
    exit()

# ========================
# USER
# GET USER INFO
def get_user_info(file_name):
    try:
        with open('./setting/' + file_name) as file:
            logger.info("[COMPLETE] User File 확인")
            info = file.readlines()
        user_info_list = info[0:2]
        return [info.strip() for info in user_info_list]
    except FileNotFoundError:
        logger.error("[ERROR] User File 확인 실패")
        exit()

# ========================
# DRUG LIST
# CHECK DRUG LIST FILE
def check_drug_list_file(file):
    try:
        drug_file = load_workbook(file)
        logger.info("[COMPLETE] Drug File 확인")
        drug_sheet = drug_file.worksheets[0]
        return drug_sheet.rows

    except FileNotFoundError:
        logger.error("[ERROR] Drug File 확인 실패")
        exit()


# ========================
# EXCEL
# MAKE EXCEL
# def make_excel(data_list, name):
#     """
#         :호출예시 make_excel([ [1,2,3,4], [5,6,7,8] ]) or make_excel(2dArray)
#         :param data_list:  [ data1, data2, data3, data4 ] 꼴의 1차원 list를 가지는 2차원 list
#         :return: 없
#     """
#     # === CONFIG
#     FILENAME = name + ".xlsx"
#
#     # === SAVE EXCEL
#     wb = Workbook()
#     ws1 = wb.worksheets[0]
#     header1 = ['댓글 아이디(완전한 이메일 형태)', '내용', '크롤링 URL', '작성시각', '현재시각']
#     ws1.column_dimensions['A'].width = 30
#     ws1.column_dimensions['B'].width = 70
#     ws1.column_dimensions['C'].width = 50
#     ws1.column_dimensions['D'].width = 20
#     ws1.column_dimensions['E'].width = 20
#     ws1.append(header1)
#
#     # DATA SAVE
#     for comment_data in data_list:
#         ws1.append(comment_data)
#     # END
#     wb.save(FILENAME)
#     #logger.info("[COMPLETE] [{}] Excel 생성 완료".format(name))


# MAKE EXCEL
def make_excel(big_info_list):
    try:
        drug_file = load_workbook("result.xlsx")
        drug_sheet = drug_file.worksheets[0]

        drug_sheet.column_dimensions['A'].width = 4
        drug_sheet.column_dimensions['B'].width = 15
        drug_sheet.column_dimensions['C'].width = 15
        drug_sheet.column_dimensions['D'].width = 50
        drug_sheet.column_dimensions['E'].width = 60
        drug_sheet.column_dimensions['F'].width = 60
        drug_sheet.column_dimensions['G'].width = 60
    except FileNotFoundError:
        drug_file = Workbook()
        drug_sheet = drug_file.worksheets[0]

    for info_list in big_info_list:
        for info in info_list:
            drug_sheet.append(info)

    drug_file.save("result.xlsx")


# SET DRUG CODE IN LIST
def set_drug_code(data_list, drug_info):
    data_list[0][0:4] = drug_info


# SET DRUG INFO IN LIST
def set_info(data_list, drug_info, idx):
    for col in range(len(drug_info)):
        data_list[col][idx] = drug_info[col]


# ========================
# HEALTH KR
def get_drug_info_heal(code):
    url_n = 'http://localapi.health.kr:8090/totalProduceN.localapi?search_word={0}&search_flag=all&sunb_count=&callback=jQuery162012134783020118656_1534985622077&_=1534985622155'.format(code)
    resp_n = requests.get(url_n).text
    data_n = resp_n.replace('jQuery162012134783020118656_1534985622077([', '').replace('])', '').strip()

    url_y = 'http://localapi.health.kr:8090/totalProduceY.localapi?search_word={0}&search_flag=all&sunb_count=&callback=jQuery162012134783020118656_1534985622077&_=1534985622155'.format(code)
    resp_y = requests.get(url_y).text
    data_y = resp_y.replace('jQuery162012134783020118656_1534985622077([', '').replace('])', '').strip()

    if data_n != '':
        return detail_heal(json.loads(data_n)['drug_code'])
    elif data_y != '':
        return detail_heal(json.loads(data_y)['drug_code'])
    else:
        return [[], []]


def detail_heal(drug_code):
    url = 'http://localapi.health.kr:8090/result_drug.localapi?drug_cd={0}&callback=jQuery16208637879955482337_1534990549348&_=1534990549525'.format(
        drug_code)
    resp = requests.get(url).text
    data = resp.replace('jQuery16208637879955482337_1534990549348([', '').replace('])', '').strip()
    mediguide = json.loads(data)['mediguide'].split('brbr')

    url2 = 'http://localapi.health.kr:8090/result_take.localapi?drug_cd={0}&callback=jQuery16205347994272901071_1535172414267&_=1535172414350'.format(
        drug_code)
    resp2 = requests.get(url2).text
    data2 = resp2.replace('jQuery16205347994272901071_1535172414267([', '').replace('])', '').strip()
    medititle = json.loads(data2)['medititle'].split('brbr')

    return [medititle, mediguide]

# ========================
# DRUGINFO
def login_drug(user_info):
    global driver

    driver.get('https://www.druginfo.co.kr')
    driver.find_element_by_xpath('/html/body/table/tbody/tr/td[3]/table/tbody/tr[6]/td/table/tbody/tr/td[3]/table/tbody/tr/td/form/table/tbody/tr[2]/td[2]/table/tbody/tr[1]/td[1]/table/tbody/tr[1]/td/label/input').send_keys(user_info[0])
    driver.find_element_by_xpath('/html/body/table/tbody/tr/td[3]/table/tbody/tr[6]/td/table/tbody/tr/td[3]/table/tbody/tr/td/form/table/tbody/tr[2]/td[2]/table/tbody/tr[1]/td[1]/table/tbody/tr[3]/td/input[1]').send_keys(user_info[1])
    driver.find_element_by_xpath('/html/body/table/tbody/tr/td[3]/table/tbody/tr[6]/td/table/tbody/tr/td[3]/table/tbody/tr/td/form/table/tbody/tr[2]/td[2]/table/tbody/tr[1]/td[2]/input').click()


def get_drug_info_drug(product_code):
    global driver

    result = list()
    # for product in product_codes:
    search_url = 'https://www.druginfo.co.kr/search2/search.aspx?q='
    url = search_url + str(product_code)
    html = requests.get(url).text
    bs = BeautifulSoup(html, 'lxml')

    try:
        trs = bs.find('div', id='main').find('div', class_='table-res').find('table').find_all('tr')
        for tr in trs:
            a_tag = tr.find('a', class_='product-link')
            if a_tag is not None:
                # result.append(detail(a_tag['href']))
                return detail_drug(a_tag['href'])
    except Exception as exc:
        # result.append('약품 없음.')
        return list()

    return result


def detail_drug(href):
    global driver

    url = 'https://www.druginfo.co.kr' + href
    driver.get(url)
    driver.implicitly_wait(3)

    time.sleep(0.2)
    bs = BeautifulSoup(driver.page_source, 'lxml')
    trs = bs.find('div', id='contents-group4').find('table').find_all('tr')
    for tr in trs:
        try:
            if '복약지도' == tr.find('td', class_='pdt-head-cell-left').get_text():
                return tr.find('td', class_='pdt-cell').get_text().strip().split('\n')
        except Exception as exc:
            continue
    return list()


# ==========================
# KIMSONLINE
# LOGIN
def login_kims(user_info):
    # MOVE TO URL
    driver.get('https://www.kimsonline.co.kr/')
    driver.implicitly_wait(3)

    # LOG IN IN KIMS ONLINE
    driver.find_element_by_xpath('//*[@id="user_id"]').send_keys(user_info[0])
    driver.find_element_by_xpath('//*[@id="user_pw"]').send_keys(user_info[1])
    driver.find_element_by_xpath('//*[@id="frmLogin"]/div[3]/div[2]/a').click()


# GET DRUG INFO IN KIMSONLINE
def get_drug_info_kims(drug_code):
    drug_data = []

    # driver.get('https://www.kimsonline.co.kr/')
    # time.sleep(0.3)
    #
    # # MOVE FOR SEARCHING DRUG
    # driver.find_element_by_xpath('//*[@id="gnb"]/li[1]/a').click()
    # driver.implicitly_wait(3)
    # time.sleep(0.2)
    #
    # # SEARCH DRUG
    # driver.find_element_by_xpath('//*[@id="txtKDCode"]').send_keys(drug_code)
    # driver.find_element_by_xpath('//*[@id="contents"]/div/div[6]/a[1]').click()
    # driver.implicitly_wait(3)
    # time.sleep(0.2)

    driver.get('http://www.kimsonline.co.kr/drugcenter/search/detailsearchlist?kdcode={}'.format(drug_code))
    driver.implicitly_wait(3)

    # SELECT DRUG
    try:
        driver.find_element_by_xpath('//*[@id="tabMarketS"]/ul/li/div[2]/div[1]/a').click()
    except:
        return []
    driver.implicitly_wait(3)

    # SELECT MENU
    driver.find_element_by_xpath('//*[@id="contents"]/div/div[5]/ul/li[4]').click()
    driver.implicitly_wait(3)
    time.sleep(0.1)

    # GET DRUG INFO
    cnt = 0
    while True:
        try:
            bs4 = BeautifulSoup(driver.page_source, 'lxml')
            drug_info = bs4.find('div', id='ctl01_area_mediguide_brief').find_all('div', class_='mt10')
            drug_data.append(drug_info[0].get_text().strip().split('\n')[0])
            for medi_guide in drug_info[1].get_text().strip().split('\n')[0:-1]:
                drug_data.append(medi_guide)
        except:
            if cnt == 5:
                break
            time.sleep(0.2)
            cnt += 1
            continue
        break
    return drug_data


def get_now_time():
    now = time.localtime()
    s = "{0}.{1:0>2}.{2:0>2}. {3:0>2}:{4:0>2}:{5:0>2}".format(now.tm_year, now.tm_mon, now.tm_mday, now.tm_hour, now.tm_min, now.tm_sec)
    return s


if __name__ == '__main__':
    # Get Setting File
    setting_list = []
    with open('./setting/setting.ini', 'r') as file:
        temp_list = file.readlines()
        setting_list.append(temp_list[1].strip())
        setting_list.append(int(temp_list[4].strip()))

    # Logger
    logger = logging.getLogger('notice')
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter('[SYSTEM] %(asctime)s :: %(message)s')
    streamHandler = logging.StreamHandler()
    streamHandler.setFormatter(formatter)
    logger.addHandler(streamHandler)

    logger.info(get_now_time())

    # Variables
    chk = 0

    # Get user info for Drug Info and Kims Online
    user_info_drug = get_user_info('user_drug.txt')
    user_info_kims = get_user_info('user_kims.txt')

    # Get drug list from drug file
    drug_rows = check_drug_list_file(setting_list[0])

    # Initiate driver
    driver = webdriver.Chrome('./setting/chromedriver.exe')
    driver.maximize_window()
    driver.implicitly_wait(3)

    # Log In in Drug Info and Kims Online
    login_drug(user_info_drug)
    login_kims(user_info_kims)

    big_data = []

    for drug_row in drug_rows:
        if chk <= setting_list[1]:
            chk += 1
            continue

        data_list = [[''] * 7 for i in range(10)]
        drug_code = drug_row[2].value

        set_drug_code(data_list, [x.value for x in drug_row])

        try:
            time1 = time.time()
            heal_list = get_drug_info_heal(drug_code)
            set_info(data_list, heal_list[0] + heal_list[1], 4)
            logger.info('heal 완료 : ' + str(time.time() - time1))
        except Exception as exc:
            logger.info('heal 오류 : ' + str(exc))

        try:
            time2 = time.time()
            set_info(data_list, get_drug_info_drug(drug_code), 5)
            time.sleep(3.0)
            logger.info('durg 완료 : ' + str(time.time() - time2))
        except Exception as exc:
            logger.info('drug 오류 : ' + str(exc))

        try:
            time3 = time.time()
            set_info(data_list, get_drug_info_kims(drug_code), 6)
            logger.info('kims 완료 : ' + str(time.time() - time3))
        except Exception as exc:
            logger.info('kims 오류 : ' + str(exc))

        big_data.append(data_list)

        if chk % 10 == 0:
            make_excel(big_data)
            big_data.clear()

        logger.info("{}번째 약품 진행 완료".format(chk))
        chk += 1

    # Quit driver
    driver.quit()

    logger.info(get_now_time())
